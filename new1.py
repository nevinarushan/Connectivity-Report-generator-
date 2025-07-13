import pdfplumber
from PIL import Image
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import os
import re
from datetime import datetime
import pytesseract
import numpy as np
import cv2
from tkinter import Tk, filedialog, messagebox
import sys


if getattr(sys, 'frozen', False):
    
    base_path = sys._MEIPASS
else:
    
    base_path = os.path.dirname(os.path.abspath(__file__))


tesseract_path = os.path.join(base_path, 'tesseract', 'tesseract.exe')
pytesseract.pytesseract.tesseract_cmd = tesseract_path


root = Tk()
root.withdraw()

messagebox.showinfo("Step 1: Select PDF Folder", "👉 Please select the folder containing the PDF reports.")
pdf_folder = filedialog.askdirectory(title="Select PDF Folder (Reports)")
if not pdf_folder:
    messagebox.showerror("Missing PDF Folder", "❌ No PDF folder selected. Exiting.")
    exit(1)

messagebox.showinfo("Step 2: Select Excel File", "👉 Please select the Excel file to update.")
excel_path = filedialog.askopenfilename(
    title="Select Excel File",
    filetypes=[("Excel files", "*.xlsx")]
)
if not excel_path:
    messagebox.showerror("Missing Excel File", "❌ No Excel file selected. Exiting.")
    exit(1)


cmbo_map = {
    "APAC STAFF": {"Packet Loss": "B", "Latency": "H", "Jitter": "N"},
    "APAC MOBILE": {"Packet Loss": "C", "Latency": "I", "Jitter": "O"},
    "SCCCL GUEST": {"Packet Loss": "D", "Latency": "J", "Jitter": "P"},
    "VLAN 82": {"Packet Loss": "E", "Latency": "K", "Jitter": "Q"},
    "VLAN 91": {"Packet Loss": "F", "Latency": "L", "Jitter": "R"},
    "VLAN 92": {"Packet Loss": "G", "Latency": "M", "Jitter": "S"},
}


pcwk_map = {
    "APAC STAFF": {"Packet Loss": "W", "Latency": "AA", "Jitter": "AE"},
    "APAC MOBILE": {"Packet Loss": "X", "Latency": "AB", "Jitter": "AF"},
    "SCCCL GUEST": {"Packet Loss": "Y", "Latency": "AC", "Jitter": "AG"},
    "VLAN": {"Packet Loss": "Z", "Latency": "AD", "Jitter": "AH"},
}


patterns = {
    "Packet Loss": r'Packet\s*loss.*?([\d\.]+%)\,',
    "Latency": r'Latency.*?([\d\.]+\s*ms)\,',
    "Jitter": r'Jitter.*?([\d\.]+\s*ms)\,',
}


def find_next_available_row(sheet, col="A"):
    row = 1
    while True:
        cell = sheet[f"{col}{row}"]
        if not isinstance(cell, MergedCell) and (cell.value is None or str(cell.value).strip() == ""):
            return row
        row += 1


wb = load_workbook(excel_path)
if "Connectivity" not in wb.sheetnames:
    messagebox.showerror("Missing Sheet", f"⚠ 'Connectivity' sheet not found in {excel_path}. Exiting.")
    exit(1)

ws = wb["Connectivity"]


now = datetime.now().strftime("%d-%m-%Y") + " - 11.00AM"
next_row = find_next_available_row(ws)
ws[f"A{next_row}"] = now  
ws[f"V{next_row}"] = now  


for filename in os.listdir(pdf_folder):
    if not filename.lower().endswith(".pdf"):
        continue

    name = os.path.splitext(filename)[0].strip().upper()
    file_path = os.path.join(pdf_folder, filename)

    with pdfplumber.open(file_path) as pdf:
        if len(pdf.pages) < 3:
            print(f"⚠ Skipping {filename}: Less than 3 pages.")
            continue
        text = pdf.pages[2].extract_text()

        if not text or text.strip() == "":
            print(f"🔍 OCR fallback for: {filename}")
            image = pdf.pages[2].to_image(resolution=300).original
            open_cv_image = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
            gray = cv2.cvtColor(open_cv_image, cv2.COLOR_BGR2GRAY)
            _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
            text = pytesseract.image_to_string(thresh)

    text = text.replace("\n", " ").strip()
    print(f"\n📄 Processing: {filename}")
    print("📝 Extracted text (preview):", text[:300])

    
    if name.endswith("PCWK"):
        base_name = name.replace(" PCWK", "").strip()
        if base_name == "LAN":
            base_name = "VLAN"

        if base_name not in pcwk_map:
            print(f"⚠ Skipping PCWK: Unrecognized source '{base_name}'")
            continue

        for metric, pattern in patterns.items():
            match = re.search(pattern, text, re.IGNORECASE)
            value = match.group(1).strip() if match else "N/A"
            col = pcwk_map[base_name][metric]
            ws[f"{col}{next_row}"] = value
            print(f"✅ PCWK {base_name} {metric}: {value} → {col}{next_row}")
        continue

    
    if name not in cmbo_map:
        print(f"⚠ Skipping CMBO: Unrecognized source '{name}'")
        continue

    for metric, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        value = match.group(1).strip() if match else "N/A"
        col = cmbo_map[name][metric]
        ws[f"{col}{next_row}"] = value
        print(f"✅ CMBO {name} {metric}: {value} → {col}{next_row}")


wb.save(excel_path)
messagebox.showinfo("Success", f"✅ All CMBO & PCWK data written successfully to:\n{excel_path}")